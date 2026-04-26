from __future__ import annotations

import csv
import re
import tempfile
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


ISTAT_COMUNI_XLSX_URL = (
    "https://www.istat.it/storage/codici-unita-amministrative/"
    "Elenco-comuni-italiani.xlsx"
)

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
OUTPUT_CSV = DATA_DIR / "comuni_italiani.csv"


TARGET_COLUMNS = [
    "comune",
    "provincia",
    "sigla",
    "regione",
]


COLUMN_ALIASES = {
    "comune": [
        "denominazione in italiano",
        "denominazione_ita",
        "denominazione italiana",
        "denominazione comune",
        "denominazione del comune",
        "comune",
        "nome comune",
    ],
    "provincia": [
        "denominazione dell'unità territoriale sovracomunale",
        "denominazione unita territoriale sovracomunale",
        "denominazione provincia",
        "provincia",
        "nome provincia",
        "città metropolitana",
        "citta metropolitana",
    ],
    "sigla": [
        "sigla automobilistica",
        "sigla provincia",
        "sigla",
        "targa",
    ],
    "regione": [
        "denominazione regione",
        "regione",
        "nome regione",
    ],
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ")
    text = " ".join(text.split()).strip()

    return text


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()

    replacements = {
        "à": "a",
        "è": "e",
        "é": "e",
        "ì": "i",
        "ò": "o",
        "ù": "u",
        "’": "'",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"[^a-z0-9]+", "_", text)

    return text.strip("_")


def find_column(headers: list[str], aliases: list[str]) -> int | None:
    normalized_headers = [normalize_header(header) for header in headers]
    normalized_aliases = [normalize_header(alias) for alias in aliases]

    # Match esatto.
    for alias in normalized_aliases:
        if alias in normalized_headers:
            return normalized_headers.index(alias)

    # Match contenitivo.
    for alias in normalized_aliases:
        for idx, header in enumerate(normalized_headers):
            if alias and alias in header:
                return idx

    return None


def download_istat_xlsx() -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    response = requests.get(
        ISTAT_COMUNI_XLSX_URL,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream,*/*"
            ),
        },
        timeout=120,
    )
    response.raise_for_status()

    content = response.content

    if not content.startswith(b"PK"):
        sample = content[:500].decode("utf-8", errors="replace")
        raise RuntimeError(
            "Il download ISTAT non sembra un file XLSX valido. "
            f"Content-Type={response.headers.get('content-type')}; "
            f"Sample={sample}"
        )

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(content)
    tmp.close()

    return Path(tmp.name)


def load_rows_from_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)

    # Nel file ISTAT di solito il primo foglio contiene i dati.
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise RuntimeError("File ISTAT vuoto.")

    header_row_idx = find_header_row(rows)

    headers = [clean_text(value) for value in rows[header_row_idx]]

    col_idx = {
        target: find_column(headers, aliases)
        for target, aliases in COLUMN_ALIASES.items()
    }

    missing = [target for target, idx in col_idx.items() if idx is None]

    if missing:
        raise RuntimeError(
            "Impossibile riconoscere alcune colonne ISTAT: "
            f"{missing}. Header trovati: {headers}"
        )

    output: list[dict[str, str]] = []

    for raw_row in rows[header_row_idx + 1 :]:
        comune = clean_text(raw_row[col_idx["comune"]])
        provincia = clean_text(raw_row[col_idx["provincia"]])
        sigla = clean_text(raw_row[col_idx["sigla"]]).upper()
        regione = clean_text(raw_row[col_idx["regione"]])

        if not comune or not regione:
            continue

        # Evita righe aggregate o note.
        if comune.lower().startswith("totale"):
            continue

        if len(sigla) != 2:
            continue

        output.append(
            {
                "comune": comune,
                "provincia": provincia,
                "sigla": sigla,
                "regione": regione,
            }
        )

    # Deduplica conservativa.
    seen = set()
    deduped: list[dict[str, str]] = []

    for row in output:
        key = (
            row["comune"].lower(),
            row["sigla"],
            row["regione"].lower(),
        )

        if key in seen:
            continue

        seen.add(key)
        deduped.append(row)

    deduped.sort(key=lambda row: (row["regione"], row["sigla"], row["comune"]))

    return deduped


def find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """
    ISTAT a volte ha righe iniziali descrittive.
    Cerchiamo la riga che contiene Comune/Regione/Sigla.
    """
    for idx, row in enumerate(rows[:30]):
        normalized = [normalize_header(value) for value in row if value is not None]
        joined = " ".join(normalized)

        has_comune = "comune" in joined or "denominazione_in_italiano" in joined
        has_regione = "regione" in joined
        has_sigla = "sigla" in joined or "automobilistica" in joined

        if has_comune and has_regione and has_sigla:
            return idx

    # Fallback: prima riga.
    return 0


def write_csv(rows: list[dict[str, str]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TARGET_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    print("Download elenco comuni ISTAT...")
    xlsx_path = download_istat_xlsx()

    try:
        print(f"Parsing XLSX: {xlsx_path}")
        rows = load_rows_from_xlsx(xlsx_path)
        write_csv(rows)

        print(f"Creato: {OUTPUT_CSV}")
        print(f"Comuni esportati: {len(rows)}")

        if len(rows) < 7800:
            raise RuntimeError(
                f"Numero comuni sospetto: {len(rows)}. "
                "Controllare struttura file ISTAT."
            )

        print("Prime 5 righe:")
        for row in rows[:5]:
            print(row)

    finally:
        try:
            xlsx_path.unlink(missing_ok=True)
        except Exception:
            pass


if __name__ == "__main__":
    main()