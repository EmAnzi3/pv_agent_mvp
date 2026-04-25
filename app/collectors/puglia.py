from __future__ import annotations

import html
import json
import re
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from app.collectors.base import CollectorResult


XLSX_URL = (
    "https://dati.puglia.it/ckan/dataset/"
    "4af29dda-fdcc-4606-bf41-ad4fa3e30790/resource/"
    "1728acca-e2fc-4dcb-bb52-0fc1eaa628c2/download/via_fer.xlsx"
)


PV_SOURCES = [
    "fotovoltaico",
    "agrivoltaico",
    "agrovoltaico",
    "agro-fotovoltaico",
    "agrofotovoltaico",
    "solare",
]


class PugliaCollector:
    source_name = "puglia"
    base_url = "https://dati.puglia.it/ckan/dataset/impianti-proposti-via-fer"

    def __init__(self):
        from requests import Session

        self.session = Session()
        self.session.headers.update(
            {
                "User-Agent": "PV-Agent-MVP/0.1",
            }
        )

    def fetch(self) -> list[CollectorResult]:
        debug_base = Path("/app/reports/debug_puglia")
        debug_base.mkdir(parents=True, exist_ok=True)

        try:
            response = self.session.get(XLSX_URL, timeout=90)
            response.raise_for_status()
        except Exception as exc:
            self._write_text(debug_base / "download_error.txt", str(exc))
            return []

        self._write_bytes(debug_base / "via_fer.xlsx", response.content)

        rows = self._read_xlsx(response.content, debug_base)
        if not rows:
            self._write_json(debug_base / "rows_empty.json", {"note": "Nessuna riga letta dal file XLSX"})
            return []

        self._write_json(debug_base / "columns.json", list(rows[0].keys()))
        self._write_json(debug_base / "sample_rows.json", rows[:30])

        results: list[CollectorResult] = []
        matched_rows: list[dict] = []
        skipped_rows: list[dict] = []
        seen_ids: set[str] = set()

        for row in rows:
            fonte = self._clean_text(row.get("fonte") or "")
            fonte_norm = self._normalize_for_match(fonte)

            if not any(src in fonte_norm for src in PV_SOURCES):
                skipped_rows.append(row)
                continue

            provincia = self._clean_text(row.get("provincia") or "")
            comune = self._clean_text(row.get("comune") or "")
            stato = self._clean_text(row.get("stato_del_procedimento") or "")
            potenza_raw = self._clean_text(row.get("potenza_mw") or "")

            if not comune and not provincia:
                continue

            power_mw = self._normalize_power_mw(potenza_raw)

            title = self._build_title(
                fonte=fonte,
                comune=comune,
                provincia=provincia,
                power_mw=power_mw,
            )

            external_id = self._build_external_id(
                provincia=provincia,
                comune=comune,
                fonte=fonte,
                power_mw=power_mw,
                stato=stato,
            )

            if external_id in seen_ids:
                continue

            seen_ids.add(external_id)
            matched_rows.append(row)

            results.append(
                CollectorResult(
                    external_id=external_id,
                    source_url=self.base_url,
                    title=title[:250],
                    payload={
                        "title": title[:500],
                        "proponent": None,
                        "status_raw": stato,
                        "region": "Puglia",
                        "province": provincia,
                        "municipalities": [comune] if comune else [],
                        "power": f"{power_mw} MW" if power_mw else None,
                        "project_type_hint": f"Puglia VIA FER - {fonte}" if fonte else "Puglia VIA FER",
                    },
                )
            )

        self._write_json(debug_base / "matched_rows_sample.json", matched_rows[:80])
        self._write_json(debug_base / "skipped_rows_sample.json", skipped_rows[:80])
        self._write_json(
            debug_base / "summary.json",
            {
                "rows_total": len(rows),
                "matched_rows": len(matched_rows),
                "skipped_rows": len(skipped_rows),
                "results": len(results),
            },
        )

        return results

    def _read_xlsx(self, content: bytes, debug_base: Path) -> list[dict]:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active

            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw:
                return []

            header_index = 0
            headers = [self._normalize_header(v) for v in rows_raw[header_index]]

            self._write_json(
                debug_base / "xlsx_debug.json",
                {
                    "sheet": ws.title,
                    "header_index": header_index,
                    "headers": headers,
                    "first_rows": [list(r) for r in rows_raw[:10]],
                },
            )

            out: list[dict] = []

            for raw_row in rows_raw[header_index + 1:]:
                if not raw_row:
                    continue

                row: dict = {}
                non_empty = False

                for idx, header in enumerate(headers):
                    if not header:
                        continue

                    value = raw_row[idx] if idx < len(raw_row) else None
                    cleaned = self._clean_text(str(value)) if value is not None else ""

                    if cleaned:
                        non_empty = True

                    row[header] = cleaned

                if non_empty:
                    out.append(row)

            return out

        except Exception as exc:
            self._write_text(debug_base / "xlsx_parse_error.txt", str(exc))
            return []

    def _build_title(
        self,
        fonte: str,
        comune: str,
        provincia: str,
        power_mw: str | None,
    ) -> str:
        fonte_clean = fonte or "FER"
        area = comune or "Comune non indicato"

        if provincia:
            area = f"{area} ({provincia})"

        if power_mw:
            return f"Impianto FER {fonte_clean} - Comune di {area} - {power_mw} MW"

        return f"Impianto FER {fonte_clean} - Comune di {area}"

    def _build_external_id(
        self,
        provincia: str,
        comune: str,
        fonte: str,
        power_mw: str | None,
        stato: str,
    ) -> str:
        base = f"{provincia}|{comune}|{fonte}|{power_mw or ''}|{stato}".lower()
        base = re.sub(r"\s+", "-", base)
        base = re.sub(r"[^a-z0-9:/._|-]", "", base)
        return base[:250]

    def _normalize_power_mw(self, value: str | None) -> str | None:
        if not value:
            return None

        cleaned = self._clean_text(value)

        # Il dataset Puglia dà già la potenza in MW.
        # Gestisce sia formato italiano "319,11" sia formato inglese "319.11".
        if "," in cleaned and "." in cleaned:
            # formato tipo 1.234,56
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            # formato tipo 319,11
            cleaned = cleaned.replace(",", ".")
        else:
            # formato tipo 319.11 oppure 319
            cleaned = cleaned

        try:
            number = float(cleaned)
        except ValueError:
            return None

        if number.is_integer():
            return str(int(number))

        return str(round(number, 6))

    def _normalize_header(self, value) -> str | None:
        if value is None:
            return None

        text = self._clean_text(str(value)).lower()
        text = html.unescape(text)
        text = text.replace("\n", " ")
        text = re.sub(r"\s+", " ", text).strip()
        text = text.replace("à", "a")
        text = text.replace("è", "e")
        text = text.replace("é", "e")
        text = text.replace("ì", "i")
        text = text.replace("ò", "o")
        text = text.replace("ù", "u")
        text = re.sub(r"[^a-z0-9]+", "_", text)
        text = text.strip("_")

        if text == "potenza_mw":
            return "potenza_mw"

        return text or None

    def _normalize_for_match(self, value: str) -> str:
        value = self._clean_text(value).lower()
        value = html.unescape(value)
        value = value.replace("à", "a")
        value = value.replace("è", "e")
        value = value.replace("é", "e")
        value = value.replace("ì", "i")
        value = value.replace("ò", "o")
        value = value.replace("ù", "u")
        return value

    def _clean_text(self, value: str) -> str:
        value = html.unescape(value or "")
        return " ".join(value.replace("\xa0", " ").split()).strip()

    def _write_text(self, path: Path, content: str) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")

    def _write_bytes(self, path: Path, content: bytes) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)

    def _write_json(self, path: Path, obj) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")