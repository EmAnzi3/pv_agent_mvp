from __future__ import annotations

import io
import re
import time
import warnings
from datetime import datetime
from urllib.parse import urlencode

from openpyxl import load_workbook

from app.collectors.base import BaseCollector, CollectorResult


BASE_URL = "https://dati.terna.it"
DOWNLOAD_CENTER_URL = "https://dati.terna.it/download-center"
EXCEL_EXPORT_URL = "https://dati.terna.it/api/sitecore/dati/downloadcenter/records"

REQUEST_SLEEP_SECONDS = 0.25

# Terna Econnextion / Download Center
DATASET = "FER"
VIEW_BY = "Region"
DB = "enti"

# Terna espone il fotovoltaico come "Solare".
TARGET_SOURCES = {"solare"}

# Page size reale intercettata dal portale.
# Il Download Center mostrava Count = 203 per 2026/04.
DEFAULT_PAGE_SIZE = 203


class TernaEconnextionCollector(BaseCollector):
    source_name = "terna_econnextion"
    base_url = BASE_URL

    def fetch(self) -> list[CollectorResult]:
        """
        Terna Econnextion espone dati aggregati, non singoli progetti.

        Fonte reale usata:
        GET /api/sitecore/dati/downloadcenter/records?f=xlsx&...

        Colonne Excel:
        - Regione
        - Tipo Impianto
        - Fonte
        - Stato Connessione
        - Potenza (MW)
        - Numero Pratiche

        Per il nostro DB questa fonte va trattata come market intelligence:
        aggregato per regione / fonte / stato connessione.
        """
        self._bootstrap_session()

        year_month = self._find_latest_available_year_month()

        if year_month is None:
            return []

        year, month = year_month

        xlsx_bytes = self._download_excel(
            year=year,
            month=month,
            page_size=DEFAULT_PAGE_SIZE,
        )

        rows = self._rows_from_excel(xlsx_bytes)

        results: list[CollectorResult] = []
        seen: set[str] = set()

        for row in rows:
            normalized = self._normalize_row(row=row, year=year, month=month)

            if normalized is None:
                continue

            external_id = normalized["external_id"]

            if external_id in seen:
                continue

            seen.add(external_id)

            results.append(
                CollectorResult(
                    external_id=external_id,
                    source_url=normalized["source_url"],
                    title=normalized["title"],
                    payload={
                        "title": normalized["title"],
                        "proponent": "Terna - Econnextion",
                        "status_raw": normalized["status_raw"],
                        "region": normalized["region"],
                        "province": None,
                        "municipalities": [],
                        "power": normalized["power"],
                        "power_mw": normalized["power_mw"],
                        "project_type_hint": "Fotovoltaico",
                        "procedure": "Richieste di connessione Terna - dato aggregato",
                        "category": "Econnextion",
                        "dataset": DATASET,
                        "view_by": VIEW_BY,
                        "filter_year": year,
                        "filter_month": month,
                        "fonte": normalized["fonte"],
                        "tipo_impianto": normalized["tipo_impianto"],
                        "stato_connessione": normalized["status_raw"],
                        "numero_pratiche": normalized["numero_pratiche"],
                        "excel_export_url": self._build_excel_url(
                            year=year,
                            month=month,
                            page_size=DEFAULT_PAGE_SIZE,
                        ),
                        "is_aggregated_market_intelligence": True,
                    },
                )
            )

        return results

    # ------------------------------------------------------------------
    # HTTP
    # ------------------------------------------------------------------

    def _headers(self, referer: str = DOWNLOAD_CENTER_URL) -> dict[str, str]:
        return {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/147.0.0.0 Safari/537.36"
            ),
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Accept-Language": "it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7",
            "Referer": referer,
            "Origin": BASE_URL,
        }

    def _bootstrap_session(self) -> None:
        """
        Apre prima il Download Center per ottenere cookie/sessione.
        Il download Excel funziona anche con requests, ma il bootstrap rende la chiamata piÃ¹ stabile.
        """
        response = self.session.get(
            DOWNLOAD_CENTER_URL,
            headers={
                "User-Agent": self._headers()["User-Agent"],
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": self._headers()["Accept-Language"],
            },
            timeout=90,
            allow_redirects=True,
        )
        response.raise_for_status()

    def _download_excel(self, year: int, month: int, page_size: int = DEFAULT_PAGE_SIZE) -> bytes:
        url = self._build_excel_url(year=year, month=month, page_size=page_size)

        response = self.session.get(
            url,
            headers=self._headers(),
            timeout=90,
            allow_redirects=True,
        )
        response.raise_for_status()

        content = response.content

        # XLSX = zip, quindi magic bytes PK.
        if not content.startswith(b"PK"):
            sample = content[:1000].decode("utf-8", errors="replace")
            sample = " ".join(sample.split())
            raise RuntimeError(
                f"Terna Excel export did not return a valid XLSX. "
                f"Status={response.status_code}; "
                f"Content-Type={response.headers.get('content-type')}; "
                f"Sample={sample[:500]}"
            )

        return content

    # ------------------------------------------------------------------
    # DISCOVERY
    # ------------------------------------------------------------------

    def _candidate_year_months(self, lookback_months: int = 8) -> list[tuple[int, int]]:
        """
        Prova mese corrente e poi mesi precedenti.
        Econnextion viene aggiornato mensilmente, ma puÃ² non essere disponibile subito.
        """
        today = datetime.now()
        year = today.year
        month = today.month

        candidates: list[tuple[int, int]] = []

        for _ in range(lookback_months):
            candidates.append((year, month))

            month -= 1
            if month == 0:
                month = 12
                year -= 1

        return candidates

    def _find_latest_available_year_month(self) -> tuple[int, int] | None:
        for year, month in self._candidate_year_months():
            try:
                xlsx_bytes = self._download_excel(
                    year=year,
                    month=month,
                    page_size=DEFAULT_PAGE_SIZE,
                )
                rows = self._rows_from_excel(xlsx_bytes)

                if rows:
                    return year, month

            except Exception:
                pass

            time.sleep(REQUEST_SLEEP_SECONDS)

        return None

    # ------------------------------------------------------------------
    # EXCEL PARSING
    # ------------------------------------------------------------------

    def _rows_from_excel(self, content: bytes) -> list[dict]:
        """
        Legge l'XLSX Terna.

        Terna genera un file Excel senza stile predefinito.
        openpyxl lo segnala con un UserWarning innocuo:
        "Workbook contains no default style, apply openpyxl's default".

        Il warning viene soppresso qui perchÃ© non indica un problema sui dati.
        """
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Workbook contains no default style.*",
                category=UserWarning,
            )

            workbook = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True,
            )

        try:
            sheet = workbook[workbook.sheetnames[0]]

            rows_iter = sheet.iter_rows(values_only=True)

            try:
                headers_raw = next(rows_iter)
            except StopIteration:
                return []

            headers = [self._clean_text(value) for value in headers_raw]

            rows: list[dict] = []

            for raw_row in rows_iter:
                if raw_row is None:
                    continue

                row: dict = {}

                for idx, header in enumerate(headers):
                    if not header:
                        continue

                    row[header] = raw_row[idx] if idx < len(raw_row) else None

                if not any(value is not None and str(value).strip() for value in row.values()):
                    continue

                rows.append(row)

            return rows

        finally:
            workbook.close()

    def _normalize_row(self, row: dict, year: int, month: int) -> dict | None:
        region = self._clean_text(row.get("Regione"))
        tipo_impianto = self._clean_text(row.get("Tipo Impianto"))
        fonte = self._clean_text(row.get("Fonte"))
        status_raw = self._clean_text(row.get("Stato Connessione"))

        power_mw = self._to_float(row.get("Potenza (MW)"))
        numero_pratiche = self._to_int(row.get("Numero Pratiche"))

        if not region:
            return None

        if not fonte:
            return None

        if fonte.lower() not in TARGET_SOURCES:
            return None

        if power_mw is None:
            return None

        if numero_pratiche is None:
            numero_pratiche = 0

        region_norm = self._normalize_key(region)
        fonte_norm = self._normalize_key(fonte)
        status_norm = self._normalize_key(status_raw or "nd")

        external_id = (
            f"terna_econnextion_{year}_{month:02d}_"
            f"{region_norm}_{fonte_norm}_{status_norm}"
        )

        title = (
            f"Terna Econnextion - {region.title()} - "
            f"{fonte} - {status_raw or 'Stato non disponibile'}"
        )

        return {
            "external_id": external_id,
            "source_url": f"{DOWNLOAD_CENTER_URL}#{external_id}",
            "title": title[:250],
            "region": region.title(),
            "tipo_impianto": tipo_impianto or DATASET,
            "fonte": fonte,
            "status_raw": status_raw,
            "power_mw": power_mw,
            "power": f"{power_mw:.6f} MW",
            "numero_pratiche": numero_pratiche,
        }

    # ------------------------------------------------------------------
    # EXPORT URL
    # ------------------------------------------------------------------

    def _build_excel_url(self, year: int, month: int, page_size: int = DEFAULT_PAGE_SIZE) -> str:
        params = {
            "f": "xlsx",
            "filterDataset": DATASET,
            "filterViewBy": VIEW_BY,
            "filterYear": str(year),
            "filterMonth": str(month),
            "orderByColumn": "Potenza (MW)",
            "orderByDir": "desc",
            "db": DB,
            "pageSize": str(page_size),
        }

        return f"{EXCEL_EXPORT_URL}?{urlencode(params)}"

    # ------------------------------------------------------------------
    # HELPERS
    # ------------------------------------------------------------------

    def _clean_text(self, value) -> str | None:
        if value is None:
            return None

        text = " ".join(str(value).replace("\xa0", " ").split()).strip()
        return text or None

    def _normalize_key(self, value: str | None) -> str:
        value = self._clean_text(value) or ""
        value = value.lower()
        value = value.replace("Ã ", "a")
        value = value.replace("Ã¨", "e")
        value = value.replace("Ã©", "e")
        value = value.replace("Ã¬", "i")
        value = value.replace("Ã²", "o")
        value = value.replace("Ã¹", "u")
        value = re.sub(r"[^a-z0-9]+", "_", value)
        value = value.strip("_")
        return value or "nd"

    def _to_float(self, value) -> float | None:
        if value is None:
            return None

        if isinstance(value, float):
            return value

        if isinstance(value, int):
            return float(value)

        text = str(value).strip()

        if not text:
            return None

        text = text.replace(" ", "")

        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")

        try:
            return float(text)
        except ValueError:
            return None

    def _to_int(self, value) -> int | None:
        if value is None:
            return None

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            return int(value)

        text = str(value).strip()

        if not text:
            return None

        text = text.replace(".", "").replace(",", "")

        try:
            return int(text)
        except ValueError:
            return None


if __name__ == "__main__":
    collector = TernaEconnextionCollector()
    items = collector.fetch()

    print(f"items: {len(items)}")

    for item in items[:40]:
        print(
            item.external_id,
            "|",
            item.title,
            "|",
            item.payload.get("region"),
            "|",
            item.payload.get("power_mw"),
            "|",
            item.payload.get("numero_pratiche"),
            "|",
            item.source_url,
        )
