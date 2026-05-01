from __future__ import annotations

import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.models import ProjectEvent, ProjectMaster


def _ensure_reports_dir() -> Path:
    path = Path(settings.reports_dir)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _safe_iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _to_float(value: Any) -> float | None:
    if value is None:
        return None

    text = str(value).strip().replace(",", ".")
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _load_json(value: str | None) -> dict:
    if not value:
        return {}

    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


class ReportBuilder:
    def __init__(self, db: Session) -> None:
        self.db = db

    def build_daily_reports(self) -> list[Path]:
        reports_dir = _ensure_reports_dir()
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        files: list[Path] = []

        files.append(self._build_new_projects_report(reports_dir / f"new_projects_{stamp}.csv"))
        files.append(self._build_status_changes_report(reports_dir / f"status_changes_{stamp}.csv"))
        files.append(self._build_snapshot_report(reports_dir / f"projects_snapshot_{stamp}.csv"))
        files.append(self._build_excel_export(reports_dir / f"pv_projects_export_{stamp}.xlsx"))

        return files

    # ------------------------------------------------------------------
    # CSV REPORTS
    # ------------------------------------------------------------------

    def _build_new_projects_report(self, path: Path) -> Path:
        since = datetime.utcnow() - timedelta(days=1)

        rows = self.db.scalars(
            select(ProjectEvent)
            .where(
                ProjectEvent.event_type == "new_project",
                ProjectEvent.created_at >= since,
            )
            .order_by(ProjectEvent.created_at.desc())
        ).all()

        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["project_id", "event_type", "status", "source", "source_url", "created_at"])

            for row in rows:
                writer.writerow([
                    row.project_id,
                    row.event_type,
                    row.status_normalized,
                    row.source_name,
                    row.source_url,
                    _safe_iso(row.created_at),
                ])

        return path

    def _build_status_changes_report(self, path: Path) -> Path:
        since = datetime.utcnow() - timedelta(days=1)

        rows = self.db.scalars(
            select(ProjectEvent)
            .where(
                ProjectEvent.event_type == "project_changed",
                ProjectEvent.created_at >= since,
            )
            .order_by(ProjectEvent.created_at.desc())
        ).all()

        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["project_id", "event_type", "status", "source", "source_url", "created_at"])

            for row in rows:
                writer.writerow([
                    row.project_id,
                    row.event_type,
                    row.status_normalized,
                    row.source_name,
                    row.source_url,
                    _safe_iso(row.created_at),
                ])

        return path

    def _build_snapshot_report(self, path: Path) -> Path:
        rows = self.db.scalars(
            select(ProjectMaster).order_by(
                ProjectMaster.region.asc().nullslast(),
                ProjectMaster.province.asc().nullslast(),
                ProjectMaster.power_mw.desc().nullslast(),
                ProjectMaster.project_name.asc(),
            )
        ).all()

        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "project_name",
                "proponent",
                "region",
                "province",
                "municipalities",
                "project_type",
                "power_mw",
                "status",
                "source",
                "url",
                "updated_at",
            ])

            for row in rows:
                writer.writerow([
                    row.project_name,
                    row.proponent,
                    row.region,
                    row.province,
                    row.municipalities,
                    row.project_type,
                    row.power_mw,
                    row.status_normalized,
                    row.primary_source,
                    row.primary_url,
                    _safe_iso(row.updated_at),
                ])

        return path

    # ------------------------------------------------------------------
    # EXCEL EXPORT
    # ------------------------------------------------------------------

    def _build_excel_export(self, path: Path) -> Path:
        since = datetime.utcnow() - timedelta(days=1)

        projects = self.db.scalars(
            select(ProjectMaster).order_by(
                ProjectMaster.region.asc().nullslast(),
                ProjectMaster.province.asc().nullslast(),
                ProjectMaster.power_mw.desc().nullslast(),
                ProjectMaster.project_name.asc(),
            )
        ).all()

        new_events = self.db.scalars(
            select(ProjectEvent)
            .options(joinedload(ProjectEvent.project))
            .where(
                ProjectEvent.event_type == "new_project",
                ProjectEvent.created_at >= since,
            )
            .order_by(ProjectEvent.created_at.desc())
        ).all()

        changed_events = self.db.scalars(
            select(ProjectEvent)
            .options(joinedload(ProjectEvent.project))
            .where(
                ProjectEvent.event_type == "project_changed",
                ProjectEvent.created_at >= since,
            )
            .order_by(ProjectEvent.created_at.desc())
        ).all()

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Sintesi"

        ws_projects = wb.create_sheet("Progetti")
        ws_new = wb.create_sheet("Nuovi_24h")
        ws_changes = wb.create_sheet("Variazioni_24h")

        self._fill_summary_sheet(ws_summary, projects, new_events, changed_events)
        self._fill_projects_sheet(ws_projects, projects)
        self._fill_events_sheet(ws_new, new_events, include_details=False)
        self._fill_events_sheet(ws_changes, changed_events, include_details=True)

        for ws in wb.worksheets:
            self._style_sheet(ws)

        wb.save(path)
        return path

    def _fill_summary_sheet(
        self,
        ws,
        projects: list[ProjectMaster],
        new_events: list[ProjectEvent],
        changed_events: list[ProjectEvent],
    ) -> None:
        total_projects = len(projects)
        total_power = sum((_to_float(p.power_mw) or 0) for p in projects)

        by_region: dict[str, dict[str, float | int]] = {}

        for project in projects:
            region = project.region or "Non indicata"
            power = _to_float(project.power_mw) or 0

            if region not in by_region:
                by_region[region] = {
                    "projects": 0,
                    "power_mw": 0.0,
                }

            by_region[region]["projects"] += 1
            by_region[region]["power_mw"] += power

        ws.append(["Indicatore", "Valore"])
        ws.append(["Totale progetti", total_projects])
        ws.append(["Potenza totale MW", round(total_power, 3)])
        ws.append(["Nuovi progetti ultime 24h", len(new_events)])
        ws.append(["Variazioni ultime 24h", len(changed_events)])
        ws.append([])
        ws.append(["Regione", "Numero progetti", "MW totali"])

        for region, values in sorted(by_region.items(), key=lambda item: str(item[0])):
            ws.append([
                region,
                values["projects"],
                round(float(values["power_mw"]), 3),
            ])

    def _fill_projects_sheet(self, ws, projects: list[ProjectMaster]) -> None:
        headers = [
            "ID",
            "Nome progetto",
            "Proponente",
            "Regione",
            "Provincia",
            "Comuni",
            "Tipo progetto",
            "Potenza MW",
            "Stato",
            "Fonte",
            "URL",
            "Aggiornato il",
        ]

        ws.append(headers)

        for project in projects:
            ws.append([
                project.id,
                project.project_name,
                project.proponent,
                project.region,
                project.province,
                project.municipalities,
                project.project_type,
                _to_float(project.power_mw),
                project.status_normalized,
                project.primary_source,
                project.primary_url,
                _safe_iso(project.updated_at),
            ])

    def _fill_events_sheet(
        self,
        ws,
        events: list[ProjectEvent],
        include_details: bool,
    ) -> None:
        headers = [
            "Evento ID",
            "Project ID",
            "Nome progetto",
            "Proponente",
            "Regione",
            "Provincia",
            "Comuni",
            "Potenza MW",
            "Stato",
            "Fonte",
            "URL",
            "Creato il",
        ]

        if include_details:
            headers.append("Dettagli variazione")

        ws.append(headers)

        for event in events:
            project = event.project
            details = _load_json(event.details_json)

            row = [
                event.id,
                event.project_id,
                project.project_name if project else None,
                project.proponent if project else None,
                project.region if project else None,
                project.province if project else None,
                project.municipalities if project else None,
                _to_float(project.power_mw) if project else None,
                event.status_normalized,
                event.source_name,
                event.source_url,
                _safe_iso(event.created_at),
            ]

            if include_details:
                row.append(json.dumps(details, ensure_ascii=False))

            ws.append(row)

    # ------------------------------------------------------------------
    # EXCEL STYLE
    # ------------------------------------------------------------------

    def _style_sheet(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_fill = PatternFill("solid", fgColor="D9EAF7")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = thin_fill

        self._format_urls(ws)
        self._format_numbers(ws)
        self._auto_width(ws)

    def _format_urls(self, ws) -> None:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    def _format_numbers(self, ws) -> None:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value

                if header and "MW" in str(header):
                    cell.number_format = "0.000"

    def _auto_width(self, ws) -> None:
        max_width = 60

        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            values = []

            for cell in column_cells:
                if cell.value is not None:
                    values.append(str(cell.value))

            if not values:
                continue

            width = min(max(len(value) for value in values) + 2, max_width)
            ws.column_dimensions[column_letter].width = width