import json
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


ROOT = Path(".")
DATA_PATH = ROOT / "docs" / "data.json"
OUT_DIR = ROOT / "reports"
OUT_DIR.mkdir(exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUT_PATH = OUT_DIR / f"proponenti_pipeline_{timestamp}.xlsx"


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, list):
        value = ", ".join(str(x) for x in value if x)
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_key(value):
    s = clean_text(value).upper()
    s = re.sub(r"\bS\.?\s*R\.?\s*L\.?\b", "SRL", s)
    s = re.sub(r"\bS\.?\s*P\.?\s*A\.?\b", "SPA", s)
    s = re.sub(r"\bSOCIETA'\b|\bSOCIETA\b", "SOCIETA", s)
    s = re.sub(r"[\"“”']", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_bad_proponent(value):
    s = clean_text(value)
    if not s:
        return True

    low = s.lower()

    bad = {
        "n/d",
        "nd",
        "none",
        "null",
        "-",
        "non disponibile",
        "proponente n/d",
    }

    if low in bad:
        return True

    # Esclude date finite per errore nel campo proponente
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", s):
        return True
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", s):
        return True

    return False


def parse_mw(value):
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    is_kw = bool(re.search(r"\bkwp?\b", raw, re.I)) and not bool(re.search(r"\bmwp?\b", raw, re.I))

    s = raw.replace("MWp", "").replace("MW", "").replace("Mwp", "")
    s = re.sub(r"\bkwp?\b", "", s, flags=re.I)
    s = re.sub(r"[^\d,.\-]", "", s)

    if not s:
        return 0.0

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        n = float(s)
    except ValueError:
        return 0.0

    return n / 1000 if is_kw else n


def get_first(record, keys, fallback=""):
    for key in keys:
        value = record.get(key)
        if value is not None and clean_text(value) != "":
            return value
    return fallback


def source_label(record):
    source = clean_text(get_first(record, ["source_group", "source", "primary_source"])).lower()
    label = clean_text(get_first(record, ["source_label", "source", "primary_source"]))

    if "terna" in source or "terna" in label.lower():
        return "Terna Econnextion"
    if source in {"mase", "mase_provvedimenti"} or "mase" in label.lower():
        return "MASE"
    if source == "sistema_puglia_energia":
        return "Sistema Puglia Energia"

    return label or "n/d"


def is_terna(record):
    return "terna" in source_label(record).lower()


def autosize(ws, max_widths=None):
    max_widths = max_widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            value = clean_text(cell.value)
            if value:
                width = max(width, min(len(value) + 2, max_widths.get(letter, 42)))
        ws.column_dimensions[letter].width = width


data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
records = data.get("records", [])

detail_rows = []
groups = {}

for r in records:
    if is_terna(r):
        continue

    proponent = clean_text(get_first(r, ["proponent", "proponente", "company"]))
    if is_bad_proponent(proponent):
        continue

    key = norm_key(proponent)

    project_name = clean_text(get_first(r, ["project_name", "title", "project", "name", "progetto"], "Progetto n/d"))
    region = clean_text(get_first(r, ["region", "regione"], ""))
    province = clean_text(get_first(r, ["province", "provincia"], ""))
    municipalities = clean_text(get_first(r, ["municipalities", "comuni", "municipality", "comune"], ""))
    mw = parse_mw(get_first(r, ["power_mw", "mw", "power"], 0))
    src = source_label(r)
    ptype = clean_text(get_first(r, ["project_type", "type", "tipo"], ""))
    status = clean_text(get_first(r, ["status_normalized", "status", "stato"], ""))
    url = clean_text(get_first(r, ["primary_url", "source_url", "url", "link"], ""))

    project_id = clean_text(get_first(r, ["project_key", "id", "external_id"], project_name + region + province + municipalities))

    if key not in groups:
        groups[key] = {
            "proponente": proponent,
            "projects": set(),
            "mw_total": 0.0,
            "mw_values": [],
            "regions": set(),
            "provinces": set(),
            "municipalities": set(),
            "sources": set(),
            "types": set(),
            "statuses": set(),
        }

    g = groups[key]
    g["projects"].add(project_id)
    g["mw_total"] += mw
    if mw > 0:
        g["mw_values"].append(mw)
    if region:
        g["regions"].add(region)
    if province:
        g["provinces"].add(province)
    if municipalities:
        for c in municipalities.split(","):
            c = clean_text(c)
            if c:
                g["municipalities"].add(c)
    if src:
        g["sources"].add(src)
    if ptype:
        g["types"].add(ptype)
    if status:
        g["statuses"].add(status)

    detail_rows.append([
        proponent,
        project_name,
        region,
        province,
        municipalities,
        mw,
        src,
        ptype,
        status,
        url,
    ])


summary_rows = []

for g in groups.values():
    n_projects = len(g["projects"])
    mw_total = g["mw_total"]
    mw_avg = mw_total / n_projects if n_projects else 0
    mw_max = max(g["mw_values"]) if g["mw_values"] else 0

    summary_rows.append([
        g["proponente"],
        n_projects,
        mw_total,
        mw_avg,
        mw_max,
        ", ".join(sorted(g["regions"])),
        ", ".join(sorted(g["provinces"])),
        ", ".join(sorted(g["municipalities"]))[:1000],
        ", ".join(sorted(g["sources"])),
        ", ".join(sorted(g["types"]))[:1000],
        ", ".join(sorted(g["statuses"]))[:1000],
    ])

summary_rows.sort(key=lambda x: (x[1], x[2]), reverse=True)
detail_rows.sort(key=lambda x: (x[0].upper(), -x[5]))


wb = Workbook()
ws = wb.active
ws.title = "Proponenti"

ws2 = wb.create_sheet("Dettaglio progetti")
ws3 = wb.create_sheet("Top 20 MW")

header_fill = PatternFill("solid", fgColor="0F766E")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2E7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

summary_headers = [
    "Proponente",
    "N. progetti",
    "MW totali",
    "MW medio",
    "MW massimo",
    "Regioni",
    "Province",
    "Comuni",
    "Fonti",
    "Tipi",
    "Stati",
]

ws.append(summary_headers)
for row in summary_rows:
    ws.append(row)

detail_headers = [
    "Proponente",
    "Progetto",
    "Regione",
    "Provincia",
    "Comune/i",
    "MW",
    "Fonte",
    "Tipo",
    "Stato",
    "Link",
]

ws2.append(detail_headers)
for row in detail_rows:
    ws2.append(row)

top_headers = ["Proponente", "N. progetti", "MW totali"]
ws3.append(top_headers)
for row in summary_rows[:20]:
    ws3.append([row[0], row[1], row[2]])


for sheet in [ws, ws2, ws3]:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.column in [3, 4, 5, 6] and sheet.title == "Proponenti":
                cell.number_format = '#,##0.000'
            if cell.column == 6 and sheet.title == "Dettaglio progetti":
                cell.number_format = '#,##0.000'
            if cell.column == 3 and sheet.title == "Top 20 MW":
                cell.number_format = '#,##0.000'


# Tabelle Excel
if ws.max_row > 1:
    tab = Table(displayName="TabellaProponenti", ref=f"A1:K{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

if ws2.max_row > 1:
    tab2 = Table(displayName="TabellaDettaglioProgetti", ref=f"A1:J{ws2.max_row}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws2.add_table(tab2)

if ws3.max_row > 1:
    tab3 = Table(displayName="TabellaTop20MW", ref=f"A1:C{ws3.max_row}")
    tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws3.add_table(tab3)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top 20 proponenti per MW"
    chart.y_axis.title = "Proponente"
    chart.x_axis.title = "MW"
    chart.height = 14
    chart.width = 24

    data_ref = Reference(ws3, min_col=3, min_row=1, max_row=ws3.max_row)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws3.add_chart(chart, "E2")


autosize(ws, {"A": 38, "H": 50, "J": 50, "K": 50})
autosize(ws2, {"A": 34, "B": 60, "E": 36, "H": 42, "J": 60})
autosize(ws3, {"A": 38})

ws.column_dimensions["A"].width = 38
ws.column_dimensions["H"].width = 50
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["J"].width = 60

for sheet in [ws, ws2]:
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = 34

wb.save(OUT_PATH)

print(f"Creato file Excel: {OUT_PATH}")
print(f"Proponenti: {len(summary_rows)}")
print(f"Progetti dettagliati: {len(detail_rows)}")
